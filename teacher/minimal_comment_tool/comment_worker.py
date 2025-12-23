#!/usr/bin/env python3
"""
Minimal Excel comment worker.
Reads JSON from stdin with keys: input_file, output_file, comment_rules
Writes comments to column M (13) based on score in column L (12) using openpyxl only.
Outputs JSON to stdout describing success or error.
"""
import sys
import json
import os
from pathlib import Path

try:
    import openpyxl
except Exception as e:
    # Print JSON error so callers can parse it
    sys.stdout.write(json.dumps({'success': False, 'message': 'Import error: ' + str(e)}))
    sys.exit(1)


def log_debug(message):
    try:
        log_dir = Path(__file__).resolve().parent.parent / 'logs'
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / 'excel_processor_minimal.log'
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] {message}\n")
    except Exception:
        pass


def get_comment(score, rules):
    try:
        s = float(score)
    except Exception:
        return ''
    for r in rules:
        try:
            if r.get('min', 0) <= s <= r.get('max', 0):
                return r.get('comment', '')
        except Exception:
            continue
    return ''


def process(input_file, output_file, rules):
    wb = openpyxl.load_workbook(input_file, data_only=True)
    log_debug(f"Loaded workbook: {input_file}. Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        start_row = 8
        log_debug(f"Processing sheet: {sheet_name}, max_row={max_row}")
        processed = 0
        commented = 0
        for row in range(start_row, max_row + 1):
            try:
                stt_cell = ws.cell(row=row, column=1)
                # Accept numeric STT or string digits (more flexible than isdigit())
                stt = ''
                if stt_cell.value is not None:
                    try:
                        # If it's numeric, convert to int string
                        if isinstance(stt_cell.value, (int, float)):
                            stt = str(int(stt_cell.value))
                        else:
                            stt = str(stt_cell.value).strip()
                    except Exception:
                        stt = str(stt_cell.value).strip()

                code_cell = ws.cell(row=row, column=2)
                student_code = str(code_cell.value).strip() if code_cell.value is not None else ''

                if not stt:
                    log_debug(f"Skip row {row} in {sheet_name}: missing STT")
                    continue
                if not (stt.isdigit() or stt.replace('.','',1).isdigit()):
                    log_debug(f"Skip row {row} in {sheet_name}: STT not numeric ({stt!r})")
                    continue
                if not student_code:
                    log_debug(f"Skip row {row} in {sheet_name}: missing student_code")
                    continue

                score_cell = ws.cell(row=row, column=12)
                score_value = score_cell.value
                if score_value is None:
                    score = 0.0
                elif isinstance(score_value, (int, float)):
                    score = float(score_value)
                else:
                    try:
                        score = float(str(score_value).replace(',', '.'))
                    except Exception:
                        score = 0.0

                comment = get_comment(score, rules)
                comment_cell = ws.cell(row=row, column=13)
                comment_cell.value = comment
                processed += 1
                if comment:
                    commented += 1
                    log_debug(f"Row {row} in {sheet_name}: STT={stt}, Code={student_code}, Score={score}, Comment='{comment}'")
                else:
                    log_debug(f"Row {row} in {sheet_name}: STT={stt}, Code={student_code}, Score={score}, no comment")
            except Exception:
                log_debug(f"Exception processing row {row} in {sheet_name}")
                continue
        log_debug(f"Sheet {sheet_name} processed: rows_with_STT={processed}, comments_written={commented}")
    # Ensure output directory exists
    outp = Path(output_file)
    outp.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(outp))
    log_debug(f"Saved processed file: {output_file}")
    return True


def main():
    try:
        input_data = json.load(sys.stdin)
    except Exception as e:
        sys.stdout.write(json.dumps({'success': False, 'message': 'Invalid JSON input: ' + str(e)}))
        sys.exit(1)

    input_file = input_data.get('input_file')
    output_file = input_data.get('output_file')
    comment_rules = input_data.get('comment_rules') or []

    # Default rules when none provided
    if not comment_rules:
        comment_rules = [
            {"min": 0,   "max": 3.5, "comment": "Chưa đạt, ý thức học kém, cần cố gắng nhiều hơn."},
            {"min": 3.5, "max": 5.0, "comment": "Chưa đạt, chưa cố gắng học, cần nghiêm túc hơn."},
            {"min": 5.0, "max": 6.5, "comment": "Có cố gắng, cần phát huy thêm."},
            {"min": 6.5, "max": 8.0, "comment": "Siêng học, cần phát huy thêm."},
            {"min": 8.0, "max": 10.0, "comment": "Chăm chỉ học tập, rất tích cực phát biểu, gương mẫu cho học sinh."}
        ]

    if not input_file or not output_file:
        sys.stdout.write(json.dumps({'success': False, 'message': 'Missing input_file or output_file'}))
        sys.exit(1)

    if not Path(input_file).exists():
        sys.stdout.write(json.dumps({'success': False, 'message': 'Input file not found: ' + str(input_file)}))
        sys.exit(1)

    try:
        ok = process(input_file, output_file, comment_rules)
        if ok:
            size = os.path.getsize(output_file)
            sys.stdout.write(json.dumps({'success': True, 'output_file': output_file, 'file_size': round(size/1024/1024, 2)}))
            sys.exit(0)
        else:
            sys.stdout.write(json.dumps({'success': False, 'message': 'Processing failed'}))
            sys.exit(2)
    except Exception as e:
        sys.stdout.write(json.dumps({'success': False, 'message': str(e)}))
        sys.exit(3)


if __name__ == '__main__':
    main()
