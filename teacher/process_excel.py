#!/usr/bin/env python3
"""
Excel Comment Processor using Pandas/OpenPyXL
Processes Excel files with multiple sheets, calculates scores from column L,
and adds comments to column M based on configurable rules.

Author: AI Assistant
Date: 2024
"""

import sys
import json
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

def log_debug(message):
    """Log debug message to file"""
    log_file = Path(__file__).parent / '..' / 'logs' / 'excel_processor.log'
    log_file.parent.mkdir(exist_ok=True)

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"[{timestamp}] {message}\n")

def get_comment(score, rules):
    """Get comment based on score and rules"""
    try:
        score = float(score)
        for rule in rules:
            if rule['min'] <= score <= rule['max']:
                return rule['comment']
        return ''
    except (ValueError, TypeError):
        return ''

def process_excel_with_pandas(input_file, output_file, comment_rules):
    """
    Process Excel file using Pandas and OpenPyXL
    """
    try:
        log_debug(f"Processing Excel file: {input_file}")

        # Handle temporary files by copying to a proper .xlsx extension if needed
        import tempfile
        import shutil

        temp_input_file = input_file
        if not input_file.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            # Create a temporary file with .xlsx extension
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_input_file = temp_file.name
            shutil.copy2(input_file, temp_input_file)
            log_debug(f"Copied temp file to: {temp_input_file}")

        # Load the Excel file with openpyxl to preserve formatting
        wb = openpyxl.load_workbook(temp_input_file, data_only=True)

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            log_debug(f"Processing sheet: {sheet_name}")

            # Find the last row with data
            max_row = ws.max_row

            # Start from row 8 (1-based indexing)
            start_row = 8

            for row in range(start_row, max_row + 1):
                try:
                    # Get STT from column A (1-based)
                    stt_cell = ws.cell(row=row, column=1)
                    stt = str(stt_cell.value).strip() if stt_cell.value is not None else ''

                    # Get student code from column B
                    code_cell = ws.cell(row=row, column=2)
                    student_code = str(code_cell.value).strip() if code_cell.value is not None else ''

                    # Skip if STT is empty, not numeric, or student code is empty
                    if not stt or not stt.isdigit() or not student_code:
                        continue

                    # Get score from column L (12th column, 1-based)
                    score_cell = ws.cell(row=row, column=12)
                    score_value = score_cell.value

                    # Handle different score formats
                    if score_value is None:
                        score = 0.0
                    elif isinstance(score_value, (int, float)):
                        score = float(score_value)
                    else:
                        # Try to parse string values
                        try:
                            score = float(str(score_value).replace(',', '.'))
                        except (ValueError, AttributeError):
                            score = 0.0

                    log_debug(f"Row {row}: STT={stt}, Code={student_code}, Score={score}")

                    # Get comment
                    comment = get_comment(score, comment_rules)

                    # Write comment to column M (13th column, 1-based)
                    comment_cell = ws.cell(row=row, column=13)
                    comment_cell.value = comment

                    log_debug(f"Added comment: {comment}")

                except Exception as e:
                    log_debug(f"Error processing row {row}: {str(e)}")
                    continue

        # Save the workbook with optimized settings
        wb.save(output_file)
        log_debug(f"Saved processed file: {output_file}")

        return True

    except Exception as e:
        log_debug(f"Error processing Excel file: {str(e)}")
        return False

def main():
    """Main processing function"""
    try:
        # Read input from stdin (JSON)
        input_data = json.load(sys.stdin)

        input_file = input_data['input_file']
        output_file = input_data['output_file']
        comment_rules = input_data['comment_rules']

        log_debug(f"Starting processing: {input_file} -> {output_file}")

        # Process the file
        success = process_excel_with_pandas(input_file, output_file, comment_rules)

        if success:
            file_size = os.path.getsize(output_file)
            result = {
                'success': True,
                'file_size': round(file_size / 1024 / 1024, 2),
                'message': 'File processed successfully with Pandas/OpenPyXL'
            }
        else:
            result = {
                'success': False,
                'message': 'Failed to process file with Pandas/OpenPyXL'
            }

        # Output result as JSON
        print(json.dumps(result))

    except Exception as e:
        log_debug(f"Main error: {str(e)}")
        result = {
            'success': False,
            'message': f'Error: {str(e)}'
        }
        print(json.dumps(result))

if __name__ == '__main__':
    main()
